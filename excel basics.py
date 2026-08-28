"""openpyxl is  a python library for reading and writing excel files .andit is automate the tasks
Installation: pip install openpyxl
check installation: pip show openpyxl
import function: from openpyxl inmport load_workbook
Workbook: A workbook is an excel file that contains one or more worksheets
 for example:"""
from openpyxl import load_workbook
wb=load_workbook("students.xlsx")
print(wb.sheetnames)

print(wb.active["J28"].value)