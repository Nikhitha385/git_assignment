from openpyxl import load_workbook
wb=load_workbook("students.xlsx")
print(wb.sheetnames)
sheet=wb["Students"]    # 
print(sheet.title)  # title of sheet
sheet=wb.active   # gets the active sheet
from openpyxl import load_workbook
wb = load_workbook("students.xlsx")
sheet = wb.active
print(sheet.title)