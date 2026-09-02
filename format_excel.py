from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Open workbook
wb = load_workbook("students.xlsx")

# Select sheets
students = wb["Students"]
summary = wb["Summary"]

# -------------------------
# COLORS / STYLES
# -------------------------

header_fill = PatternFill(
    fill_type="solid",
    fgColor="4472C4"
)

header_font = Font(
    bold=True,
    color="FFFFFF"
)

title_font = Font(
    bold=True,
    color="FFFFFF",
    size=14
)

thin = Side(style="thin")

border = Border(
    left=thin,
    right=thin,
    top=thin,
    bottom=thin
)

center = Alignment(
    horizontal="center",
    vertical="center"
)

# -------------------------
# STUDENTS SHEET
# -------------------------

# Header
for cell in students[1]:
    if cell.value is not None:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

# Find actual used columns
for row in students.iter_rows(
    min_row=2,
    max_row=students.max_row
):
    for cell in row:
        if cell.value is not None:
            cell.alignment = center
            cell.border = border

# Column widths
students.column_dimensions["A"].width = 15
students.column_dimensions["B"].width = 12
students.column_dimensions["C"].width = 12
students.column_dimensions["D"].width = 12
students.column_dimensions["E"].width = 12
students.column_dimensions["F"].width = 12
students.column_dimensions["G"].width = 4
students.column_dimensions["H"].width = 20
students.column_dimensions["I"].width = 15

# Row height
students.row_dimensions[1].height = 25

# -------------------------
# SUMMARY SHEET
# -------------------------

# Title
summary["A1"].font = title_font
summary["A1"].fill = header_fill
summary["A1"].alignment = center
summary["A1"].border = border

# Merge title
summary.merge_cells("A1:B1")

# Summary labels and values
for row in range(2, 6):

    summary[f"A{row}"].font = Font(bold=True)
    summary[f"A{row}"].alignment = Alignment(
        horizontal="left",
        vertical="center"
    )

    summary[f"B{row}"].alignment = center

    summary[f"A{row}"].border = border
    summary[f"B{row}"].border = border

# Number formatting
summary["B3"].number_format = "0.00"

# Column widths
summary.column_dimensions["A"].width = 22
summary.column_dimensions["B"].width = 15

# Row heights
summary.row_dimensions[1].height = 30

for row in range(2, 6):
    summary.row_dimensions[row].height = 22

# Save
wb.save("students.xlsx")

print("Excel alignment and formatting completed")