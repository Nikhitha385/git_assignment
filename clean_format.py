from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Open workbook
wb = load_workbook("students.xlsx")

students = wb["Students"]

# -------------------------------------------------
# 1. Remove the old statistics section
# -------------------------------------------------

for row in students.iter_rows(
    min_row=1,
    max_row=students.max_row,
    min_col=6,
    max_col=7
):
    for cell in row:
        cell.value = None
        cell._style = None

# -------------------------------------------------
# 2. Find the Marks column automatically
# -------------------------------------------------

marks_column = None

for cell in students[1]:
    if cell.value == "Marks":
        marks_column = cell.column
        break

if marks_column is None:
    print("Marks column not found")
else:

    # -------------------------------------------------
    # 3. Create statistics in H:I
    # -------------------------------------------------

    students["H1"] = "Statistics"

    last_row = students.max_row

    students["H2"] = "Total Marks"
    students["I2"] = f"=SUM({chr(64 + marks_column)}2:{chr(64 + marks_column)}{last_row})"

    students["H3"] = "Average Marks"
    students["I3"] = f"=AVERAGE({chr(64 + marks_column)}2:{chr(64 + marks_column)}{last_row})"

    students["H4"] = "Highest Marks"
    students["I4"] = f"=MAX({chr(64 + marks_column)}2:{chr(64 + marks_column)}{last_row})"

    students["H5"] = "Lowest Marks"
    students["I5"] = f"=MIN({chr(64 + marks_column)}2:{chr(64 + marks_column)}{last_row})"

    # -------------------------------------------------
    # 4. Formatting
    # -------------------------------------------------

    blue_fill = PatternFill(
        fill_type="solid",
        fgColor="4472C4"
    )

    white_bold = Font(
        bold=True,
        color="FFFFFF"
    )

    bold_font = Font(bold=True)

    thin = Side(style="thin")

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    center = Alignment(
        horizontal="center",
        vertical="center"
    )

    # Main header
    for cell in students[1]:
        if cell.value is not None:
            cell.fill = blue_fill
            cell.font = white_bold
            cell.alignment = center
            cell.border = border

    # Table formatting
    for row in students.iter_rows(
        min_row=2,
        max_row=students.max_row,
        min_col=1,
        max_col=marks_column + 2
    ):
        for cell in row:
            if cell.value is not None:
                cell.alignment = center
                cell.border = border

    # Statistics formatting
    students["H1"].fill = blue_fill
    students["H1"].font = white_bold
    students["H1"].alignment = center
    students["H1"].border = border

    for row in range(2, 6):
        students[f"H{row}"].font = bold_font
        students[f"H{row}"].border = border
        students[f"I{row}"].border = border
        students[f"I{row}"].alignment = center

    students["I3"].number_format = "0.00"

    # Column widths
    students.column_dimensions["A"].width = 15
    students.column_dimensions["B"].width = 12
    students.column_dimensions["C"].width = 12
    students.column_dimensions["D"].width = 12
    students.column_dimensions["E"].width = 12
    students.column_dimensions["F"].width = 12
    students.column_dimensions["G"].width = 4
    students.column_dimensions["H"].width = 20
    students.column_dimensions["I"].width = 15

    # Row height
    students.row_dimensions[1].height = 25

    # Save
    wb.save("students.xlsx")

    print("Workbook cleaned and formatted successfully")