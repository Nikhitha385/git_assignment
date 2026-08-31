from openpyxl import load_workbook

wb = load_workbook("students.xlsx")
sheet = wb["Students"]

highest = 0

# Step 1: Find the highest mark
for row in range(2, sheet.max_row + 1):
    marks = sheet.cell(row=row, column=2).value

    if marks > highest:
        highest = marks

# Step 2: Find all students who have that mark
print("Highest Marks:", highest)
print("Students with highest marks:")

for row in range(2, sheet.max_row + 1):
    name = sheet.cell(row=row, column=1).value
    marks = sheet.cell(row=row, column=2).value

    if marks == highest:
        print(name)