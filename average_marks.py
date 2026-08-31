from openpyxl import load_workbook

wb = load_workbook("students.xlsx")
sheet = wb["Students"]

total = 0
count = 0

for row in range(2, sheet.max_row + 1):
    marks = sheet.cell(row=row, column=2).value

    total = total + marks
    count = count + 1

average = total / count

print("Total Marks:", total)
print("Number of Students:", count)
print("Average Marks:", average)