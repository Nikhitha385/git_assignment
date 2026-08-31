from openpyxl import load_workbook

wb = load_workbook("students.xlsx")
sheet = wb["Students"]

for row in range(2, sheet.max_row + 1):
    name = sheet.cell(row=row, column=1).value
    marks = sheet.cell(row=row, column=2).value

    if marks >= 40:
        result = "Pass"
    else:
        result = "Fail"

    print(name, "-", result)