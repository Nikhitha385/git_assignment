from openpyxl import load_workbook

wb = load_workbook("students.xlsx")
sheet = wb["Students"]

lowest = float("inf")

# Step 1: Find the lowest mark
for row in range(2, sheet.max_row + 1):
    marks = sheet.cell(row=row, column=2).value

    if marks < lowest:
        lowest = marks

# Step 2: Find all students with the lowest mark
print("Lowest Marks:", lowest)
print("Students with lowest marks:")

for row in range(2, sheet.max_row + 1):
    name = sheet.cell(row=row, column=1).value
    marks = sheet.cell(row=row, column=2).value

    if marks == lowest:
        print(name)