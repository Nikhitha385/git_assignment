import os
from openpyxl import load_workbook

file_path = os.path.abspath("students.xlsx")

print("Python is using:")
print(file_path)

wb = load_workbook(file_path)
sheet = wb["Students"]

print("Number of rows:", sheet.max_row)

for row in sheet.iter_rows(values_only=True):
    print(row)