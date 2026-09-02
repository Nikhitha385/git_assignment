from openpyxl import load_workbook

# Open the Excel workbook
wb = load_workbook("students.xlsx")

# Select Students sheet
students = wb["Students"]

# Find the Marks column automatically
marks_column = None

for cell in students[1]:
    if cell.value == "Marks":
        marks_column = cell.column
        break

# Check if Marks column exists
if marks_column is None:
    print("Marks column not found.")
else:
    # Store valid student marks
    marks_list = []

    for row in range(2, students.max_row + 1):
        name = students.cell(row=row, column=1).value
        marks = students.cell(row=row, column=marks_column).value

        if name is not None and isinstance(marks, (int, float)):
            marks_list.append(marks)

    if not marks_list:
        print("No valid student marks found.")
    else:

        # Get existing Summary sheet or create it
        if "Summary" in wb.sheetnames:
            summary = wb["Summary"]
        else:
            summary = wb.create_sheet("Summary")

        # Title
        summary["A1"] = "Student Summary"

        # Total students
        total_students = len(marks_list)
        summary["A2"] = "Total Students"
        summary["B2"] = total_students

        # Average
        total_marks = sum(marks_list)
        average = total_marks / total_students

        summary["A3"] = "Average Marks"
        summary["B3"] = average

        # Highest
        highest = max(marks_list)
        summary["A4"] = "Highest Marks"
        summary["B4"] = highest

        # Lowest
        lowest = min(marks_list)
        summary["A5"] = "Lowest Marks"
        summary["B5"] = lowest

        # Save workbook
        wb.save("students.xlsx")

        print("Summary updated successfully")