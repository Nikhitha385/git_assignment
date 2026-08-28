from openpyxl import load_workbook

wb = load_workbook("students.xlsx")

sheet = wb["Students"]

for row in sheet.iter_rows(values_only=True):
    print(row)