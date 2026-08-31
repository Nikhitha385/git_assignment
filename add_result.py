from openpyxl import load_workbook

wb = load_workbook("students.xlsx")
sheet = wb["Students"]

sheet["D1"] = "Result"

for row in range(2, sheet.max_row + 1):
    marks = sheet.cell(row=row, column=2).value

    if marks >= 40:
        result = "Pass"
    else:
        result = "Fail"

    sheet.cell(row=row, column=4).value = result

wb.save("students.xlsx")

print("Results added to Excel")