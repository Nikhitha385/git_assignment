from openpyxl import load_workbook

wb = load_workbook("students.xlsx")

sheet = wb["Students"]

# Find the last row
last_row = sheet.max_row

# Add statistics heading
sheet["H1"] = "Statistics"

# Total Marks
sheet["H2"] = "Total Marks"
sheet["I2"] = f"=SUM(D2:D{last_row})"

# Average Marks
sheet["H3"] = "Average Marks"
sheet["I3"] = f"=AVERAGE(D2:D{last_row})"

# Highest Marks
sheet["H4"] = "Highest Marks"
sheet["I4"] = f"=MAX(D2:D{last_row})"

# Lowest Marks
sheet["H5"] = "Lowest Marks"
sheet["I5"] = f"=MIN(D2:D{last_row})"

wb.save("students.xlsx")

print("Formulas added successfully")